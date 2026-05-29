"""
Excel exporter for AST (After School Tennis) fee reports.

One sheet per group (TRAINER or BALLBOY), styled to match the POS Excel reports.

TRAINER columns (A–J):
  Name | Hours | Rate | Total Amount | VAT (12%) | Ex-VAT | 5% Comm | Net Amount | EWT | Final Net

BALLBOY columns (A–F):
  Name | Hours | Rate | Total Amount | VAT (12%) | Net Total
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.calculator import format_date_range

# ── Styles ────────────────────────────────────────────────────────────────────
FILL_TITLE  = PatternFill("solid", fgColor="C6EFCE")
FILL_HEADER = PatternFill("solid", fgColor="E8884C")
FILL_GT     = PatternFill("solid", fgColor="FFF3E0")
FILL_SIG    = PatternFill("solid", fgColor="E8F5E9")
FILL_ALT    = PatternFill("solid", fgColor="FFF8F4")

FONT_TITLE  = Font(name="Calibri", bold=True,  size=11)
FONT_HDR    = Font(name="Calibri", bold=True,  size=10, color="FFFFFF")
FONT_DATA   = Font(name="Calibri",             size=10)
FONT_GT     = Font(name="Calibri", bold=True,  size=10)
FONT_SIG_L  = Font(name="Calibri", bold=True,  size=9)
FONT_SIG_N  = Font(name="Calibri", bold=True,  size=10)
FONT_SIG_T  = Font(name="Calibri",             size=9)

_S = Side(style="thin", color="B0B0B0")
BORDER = Border(left=_S, right=_S, top=_S, bottom=_S)

NUM_FMT = '#,##0.00'

# ── Column specs (built dynamically from day_cols) ───────────────────────────

def _trainer_hdrs(day_labels: list) -> list:
    return (
        ["Name"]
        + day_labels
        + ["Hours", "Rate", "Total Amount", "VAT (12%)",
           "Ex-VAT", "5% Comm", "Net Amount", "EWT", "Final Net"]
    )

def _trainer_widths(n_days: int) -> list:
    return [32] + [8] * n_days + [8, 12, 18, 14, 18, 12, 18, 10, 18]

def _ballboy_hdrs(day_labels: list) -> list:
    return (
        ["Name"]
        + day_labels
        + ["Hours", "Rate", "Total Amount", "VAT (12%)", "Net Total"]
    )

def _ballboy_widths(n_days: int) -> list:
    return [32] + [8] * n_days + [8, 12, 18, 14, 18]


# ── Public API ────────────────────────────────────────────────────────────────

def export(summaries: dict, settings: dict, output_path: str) -> None:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for label, data in summaries.items():
        sheet_name = _safe_name(label)
        ws = wb.create_sheet(title=sheet_name)
        if data["type"] == "TRAINER":
            _write_trainer_sheet(ws, label, data, settings)
        else:
            _write_ballboy_sheet(ws, label, data, settings)

    wb.save(output_path)


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_trainer_sheet(ws, label, data, settings):
    day_cols  = data.get("day_cols", [])   # [(key, label), ...]
    day_labels = [lbl for (_, lbl) in day_cols]
    hdrs   = _trainer_hdrs(day_labels)
    widths = _trainer_widths(len(day_cols))
    n_cols = len(hdrs)
    last   = get_column_letter(n_cols)
    df     = data["rows"]
    date_str = data.get("date_str", "")
    title  = _make_title(label, "TRAINER")

    # Set column widths
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1: blank
    ws.append([""] * n_cols)

    # Row 2: company name (merged)
    ws.append(["Manila Polo Club Inc."] + [""] * (n_cols - 1))
    ws.merge_cells(f"A2:{last}2")
    _style_title_row(ws, 2, n_cols)

    # Row 3: report title + date
    ws.append([title] + [""] * (n_cols - 2) + [date_str])
    ws.merge_cells(f"A3:{get_column_letter(n_cols - 1)}3")
    _style_title_row(ws, 3, n_cols)
    ws.cell(3, n_cols).alignment = Alignment(horizontal="right")

    # Row 4: column headers
    ws.append(hdrs)
    _style_header_row(ws, 4, n_cols)

    # Data rows
    data_start = 5
    for i, (_, r) in enumerate(df.iterrows()):
        day_vals = [r.get(key, 0) for (key, _) in day_cols]
        row = (
            [r["name"]]
            + day_vals
            + [
                r["hours"],
                r["rate"],
                r["total_amount"],
                r["vat"],
                r["ex_vat"],
                r["commission"],
                r["net_amount"],
                r["ewt"],
                r["net_final"],
            ]
        )
        ws.append(row)
        row_n = data_start + i
        fill  = FILL_ALT if i % 2 == 0 else None
        _style_data_row(ws, row_n, n_cols, fill=fill, num_start=2)

    # Grand total row
    gt_row = data_start + len(df)
    day_sums = [
        df[key].sum() if key in df.columns and not df.empty else 0
        for (key, _) in day_cols
    ]
    ws.append(
        ["GRAND TOTAL"]
        + day_sums
        + [
            df["hours"].sum()        if not df.empty else 0,
            "",
            df["total_amount"].sum() if not df.empty else 0,
            df["vat"].sum()          if not df.empty else 0,
            df["ex_vat"].sum()       if not df.empty else 0,
            df["commission"].sum()   if not df.empty else 0,
            df["net_amount"].sum()   if not df.empty else 0,
            df["ewt"].sum()          if not df.empty else 0,
            data["grand_total"],
        ]
    )
    _style_gt_row(ws, gt_row, n_cols, num_start=2)

    # Signatories
    _write_sigs(ws, gt_row + 2, n_cols, settings)


def _write_ballboy_sheet(ws, label, data, settings):
    day_cols   = data.get("day_cols", [])   # [(key, label), ...]
    day_labels = [lbl for (_, lbl) in day_cols]
    hdrs   = _ballboy_hdrs(day_labels)
    widths = _ballboy_widths(len(day_cols))
    n_cols = len(hdrs)
    last   = get_column_letter(n_cols)
    df     = data["rows"]
    date_str = data.get("date_str", "")
    title  = _make_title(label, "BALLBOY")

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.append([""] * n_cols)

    ws.append(["Manila Polo Club Inc."] + [""] * (n_cols - 1))
    ws.merge_cells(f"A2:{last}2")
    _style_title_row(ws, 2, n_cols)

    ws.append([title] + [""] * (n_cols - 2) + [date_str])
    ws.merge_cells(f"A3:{get_column_letter(n_cols - 1)}3")
    _style_title_row(ws, 3, n_cols)
    ws.cell(3, n_cols).alignment = Alignment(horizontal="right")

    ws.append(hdrs)
    _style_header_row(ws, 4, n_cols)

    data_start = 5
    for i, (_, r) in enumerate(df.iterrows()):
        day_vals = [r.get(key, 0) for (key, _) in day_cols]
        ws.append(
            [r["name"]]
            + day_vals
            + [r["hours"], r["rate"], r["total"], r["vat"], r["net_total"]]
        )
        fill = FILL_ALT if i % 2 == 0 else None
        _style_data_row(ws, data_start + i, n_cols, fill=fill, num_start=2)

    gt_row = data_start + len(df)
    day_sums = [
        df[key].sum() if key in df.columns and not df.empty else 0
        for (key, _) in day_cols
    ]
    ws.append(
        ["GRAND TOTAL"]
        + day_sums
        + [
            df["hours"].sum() if not df.empty else 0,
            "",
            df["total"].sum() if not df.empty else 0,
            df["vat"].sum()   if not df.empty else 0,
            data["grand_total"],
        ]
    )
    _style_gt_row(ws, gt_row, n_cols, num_start=2)

    _write_sigs(ws, gt_row + 2, n_cols, settings)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _style_title_row(ws, row_n, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row_n, c)
        cell.fill   = FILL_TITLE
        cell.font   = FONT_TITLE
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_header_row(ws, row_n, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row_n, c)
        cell.fill   = FILL_HEADER
        cell.font   = FONT_HDR
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_n].height = 30


def _style_data_row(ws, row_n, n_cols, fill=None, num_start=2):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row_n, c)
        cell.font   = FONT_DATA
        cell.border = BORDER
        if fill:
            cell.fill = fill
        if c == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c == num_start - 1:
            # hours column – integer-ish
            cell.number_format = '#,##0.##'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")


def _style_gt_row(ws, row_n, n_cols, num_start=2):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row_n, c)
        cell.fill   = FILL_GT
        cell.font   = FONT_GT
        cell.border = BORDER
        if c == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c == num_start - 1:
            cell.number_format = '#,##0.##'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")


def _write_sigs(ws, start_row, n_cols, settings):
    col_w = max(n_cols // 3, 1)
    labels = ["Prepared By:", "Checked by:", "Validated by:"]
    names  = [
        settings.get("prepared_by_name",  ""),
        settings.get("checked_by_name",   ""),
        settings.get("validated_by_name", ""),
    ]
    titles = [
        settings.get("prepared_by_title",  ""),
        settings.get("checked_by_title",   ""),
        settings.get("validated_by_title", ""),
    ]
    for offset, (lbl, nm, tt) in enumerate(zip(labels, names, titles)):
        col = offset * col_w + 1
        end = col + col_w - 1

        lbl_cell = ws.cell(start_row, col, lbl)
        lbl_cell.font = FONT_SIG_L
        lbl_cell.fill = FILL_SIG
        if end > col:
            ws.merge_cells(
                start_row=start_row, start_column=col,
                end_row=start_row, end_column=end,
            )

        nm_cell = ws.cell(start_row + 2, col, nm)
        nm_cell.font = FONT_SIG_N
        if end > col:
            ws.merge_cells(
                start_row=start_row + 2, start_column=col,
                end_row=start_row + 2, end_column=end,
            )

        tt_cell = ws.cell(start_row + 3, col, tt)
        tt_cell.font = FONT_SIG_T
        if end > col:
            ws.merge_cells(
                start_row=start_row + 3, start_column=col,
                end_row=start_row + 3, end_column=end,
            )


def _make_title(label: str, grp_type: str) -> str:
    import re
    clean = re.sub(r'\s*\(?\bast\b\)?', '', label, flags=re.IGNORECASE).strip()
    kind  = "Instructor Fee" if grp_type == "TRAINER" else "Fee"
    return f"{clean.title()} (AST) {kind}"


def _safe_name(name: str) -> str:
    """Make a valid Excel sheet name (max 31 chars, no special chars)."""
    import re
    clean = re.sub(r'[\\/*?:\[\]]', '', name)[:31]
    return clean or "Sheet"
