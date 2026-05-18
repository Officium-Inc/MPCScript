"""
Generates an Excel summary workbook using openpyxl.

One sheet per Income Center Group, layout matching the sample:

  Row 1  (blank)
  Row 2  Manila Polo Club Inc.         (A2, merged A2:H2, green bg)
  Row 3  <Report Title>   <Date Range> (A3:D3 | F3:H3, green bg)
  Row 4  Column headers                (A4:H4, orange bg, bold)
  Row 5+ Data rows                     (white / light-alt)
  Last   Grand Total row               (bold, amber bg)
  +2     Signatory labels row          (green bg)
  +1     blank
  +1     Signatory names row
  +1     Signatory titles row
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from core.calculator import format_date_range, group_to_title

# ---------------------------------------------------------------------------
# Colour constants  (ARGB strings used by openpyxl)
# ---------------------------------------------------------------------------
FILL_TITLE   = PatternFill("solid", fgColor="C6EFCE")    # light green
FILL_HEADER  = PatternFill("solid", fgColor="E8884C")    # orange
FILL_GT      = PatternFill("solid", fgColor="FFF3E0")    # light amber
FILL_SIG     = PatternFill("solid", fgColor="E8F5E9")    # light green
FILL_ALT     = PatternFill("solid", fgColor="FFF8F4")    # very light pink
FILL_NONE    = PatternFill(fill_type=None)

FONT_TITLE   = Font(name="Calibri", bold=True, size=11)
FONT_HEADER  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
FONT_DATA    = Font(name="Calibri", size=10)
FONT_GT      = Font(name="Calibri", bold=True, size=10)
FONT_SIG_LBL = Font(name="Calibri", bold=True, size=9)
FONT_SIG_NM  = Font(name="Calibri", bold=True, size=10)
FONT_SIG_TL  = Font(name="Calibri", size=9)

THIN_SIDE = Side(style="thin", color="B0B0B0")
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE,
    top=THIN_SIDE, bottom=THIN_SIDE,
)

NUM_FMT = '#,##0.00'
INT_FMT = '#,##0'

# Column widths in characters
COL_WIDTHS_CHARS = [28, 16, 10, 20, 14, 16, 12, 16]

# Column headers
COL_HEADERS = [
    "Item",
    "Min of Unit Price",
    "Sum of Qty",
    "Sum of Amt. After Disc.",
    "5% Comm",
    "Total",
    "Ewt",
    "Total",
]

# Number of data columns (A–H = 8)
N_COLS = 8


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export(summaries: dict, settings: dict, output_path: str) -> None:
    """
    Write a .xlsx workbook to output_path with one sheet per group.

    Args:
        summaries:   result from calculator.summarise()
        settings:    dict from config.settings.load()
        output_path: destination file path (will be created / overwritten)
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)   # remove the default blank sheet

    for group_name, data in summaries.items():
        _write_sheet(wb, group_name, data, settings)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Sheet builder
# ---------------------------------------------------------------------------

def _write_sheet(wb: Workbook, group_name: str, data: dict, settings: dict) -> None:
    title_text  = group_to_title(group_name)
    date_range  = format_date_range(data["date_min"], data["date_max"])
    rows_df     = data["rows"]
    grand_total = data["grand_total"]

    # Sheet name: max 31 chars, strip invalid chars
    sheet_name = group_name.title()[:31]
    ws = wb.create_sheet(title=sheet_name)

    # Set column widths
    for col_idx, width in enumerate(COL_WIDTHS_CHARS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ---- Row 1: blank ----
    ws.append([])

    # ---- Row 2: Company name (merged A2:H2) ----
    ws.append(["Manila Polo Club Inc."])
    _merge_and_style(ws, 2, 1, 2, N_COLS, FILL_TITLE, FONT_TITLE, "center")

    # ---- Row 3: Report title  |  Date range ----
    row3 = [""] * N_COLS
    row3[0] = title_text
    row3[5] = date_range          # Column F
    ws.append(row3)
    _apply_fill_row(ws, 3, FILL_TITLE)
    _style_cell(ws.cell(3, 1), font=FONT_TITLE, alignment="left")
    _style_cell(ws.cell(3, 6), font=Font(name="Calibri", bold=True, size=10),
                alignment="right")
    # Merge title part A3:E3
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
    # Merge date part F3:H3
    ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=N_COLS)

    # ---- Row 4: Column headers ----
    ws.append(COL_HEADERS)
    for col_idx in range(1, N_COLS + 1):
        cell = ws.cell(4, col_idx)
        cell.fill   = FILL_HEADER
        cell.font   = FONT_HEADER
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 30

    # ---- Data rows ----
    data_start_row = 5
    for row_num, (_, row) in enumerate(rows_df.iterrows(), start=data_start_row):
        fill = FILL_ALT if row_num % 2 == 0 else FILL_NONE
        values = [
            row["item"],
            row["min_unit_price"],
            int(row["sum_qty"]),
            row["sum_amt"],
            row["comm"],
            row["total"],
            row["ewt"],
            row["final_total"],
        ]
        ws.append(values)
        _style_data_row(ws, row_num, fill)

    # ---- Grand Total row ----
    gt_row_num = data_start_row + len(rows_df)
    gt = grand_total
    gt_values = [
        gt["item"],
        "",
        gt["sum_qty"],
        gt["sum_amt"],
        gt["comm"],
        gt["total"],
        gt["ewt"],
        gt["final_total"],
    ]
    ws.append(gt_values)
    _style_grand_total_row(ws, gt_row_num)

    # ---- Spacer row ----
    ws.append([])
    spacer_row = gt_row_num + 1

    # ---- Signatory section ----
    sig_label_row = spacer_row + 1
    sig_name_row  = sig_label_row + 1
    sig_title_row = sig_label_row + 2

    # Label row: "Prepared By:" | blank | "Checked by:" | blank | blank | "Validated by:"
    sig_labels = [""] * N_COLS
    sig_labels[0] = "Prepared By:"
    sig_labels[2] = "Checked by:"
    sig_labels[5] = "Validated by:"
    ws.append(sig_labels)
    _apply_fill_row(ws, sig_label_row, FILL_SIG)
    for col_idx in [1, 3, 6]:
        ws.cell(sig_label_row, col_idx).font = FONT_SIG_LBL

    # Blank row inside signatory
    ws.append([""] * N_COLS)
    _apply_fill_row(ws, sig_label_row + 1, FILL_SIG)

    # Name row
    sig_names = [""] * N_COLS
    sig_names[0] = settings.get("prepared_by_name", "")
    sig_names[2] = settings.get("checked_by_name", "")
    sig_names[5] = settings.get("validated_by_name", "")
    ws.append(sig_names)
    _apply_fill_row(ws, sig_name_row, FILL_SIG)
    for col_idx in [1, 3, 6]:
        ws.cell(sig_name_row, col_idx).font = FONT_SIG_NM

    # Title row
    sig_titles = [""] * N_COLS
    sig_titles[0] = settings.get("prepared_by_title", "")
    sig_titles[2] = settings.get("checked_by_title", "")
    sig_titles[5] = settings.get("validated_by_title", "")
    ws.append(sig_titles)
    _apply_fill_row(ws, sig_title_row, FILL_SIG)
    for col_idx in [1, 3, 6]:
        ws.cell(sig_title_row, col_idx).font = FONT_SIG_TL

    # Box border around signatory block
    _box_range(ws, sig_label_row, 1, sig_title_row, N_COLS)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _style_cell(cell, font=None, fill=None, alignment="left"):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    h_align = {"left": "left", "center": "center", "right": "right"}.get(alignment, "left")
    cell.alignment = Alignment(horizontal=h_align, vertical="center")


def _merge_and_style(ws, row, c1, _row2, c2, fill, font, alignment="center"):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row, c1)
    cell.fill  = fill
    cell.font  = font
    cell.alignment = Alignment(horizontal=alignment, vertical="center")


def _apply_fill_row(ws, row_num, fill):
    for col_idx in range(1, N_COLS + 1):
        ws.cell(row_num, col_idx).fill = fill


def _style_data_row(ws, row_num, fill):
    for col_idx in range(1, N_COLS + 1):
        cell = ws.cell(row_num, col_idx)
        cell.font   = FONT_DATA
        cell.border = THIN_BORDER
        if fill and fill.fill_type:
            cell.fill = fill
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx == 3:
            cell.number_format = INT_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx >= 2:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")


def _style_grand_total_row(ws, row_num):
    for col_idx in range(1, N_COLS + 1):
        cell = ws.cell(row_num, col_idx)
        cell.fill   = FILL_GT
        cell.font   = FONT_GT
        cell.border = THIN_BORDER
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_idx == 3:
            cell.number_format = INT_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx >= 2:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal="right", vertical="center")


def _box_range(ws, r1, c1, r2, c2):
    """Apply a thin box border around a rectangular range."""
    for row_idx in range(r1, r2 + 1):
        for col_idx in range(c1, c2 + 1):
            cell = ws.cell(row_idx, col_idx)
            top    = THIN_SIDE if row_idx == r1 else Side()
            bottom = THIN_SIDE if row_idx == r2 else Side()
            left   = THIN_SIDE if col_idx == c1 else Side()
            right  = THIN_SIDE if col_idx == c2 else Side()
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
